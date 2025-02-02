import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import Border, Side
from openpyxl.utils import range_boundaries
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from tkinter import Tk, Label, Entry, Button, messagebox, Frame
from collections import defaultdict
import os
import sys
import re

overall_pokemon_usage = defaultdict(int)

def resource_path(relative_path):
    try:
        # PyInstaller creates a _MEIPASS folder and places resources there
        base_path = sys._MEIPASS
    except Exception:
        # If running as a normal script, use the current directory
        base_path = os.path.abspath(".")
    
    return os.path.join(base_path, relative_path)
    
sprite_folder = resource_path("regular")

def fetch_replay_data(replay_url):
    response = requests.get(replay_url)
    if response.status_code == 200:
        return response.text
    else:
        raise Exception(f"Failed to fetch replay data. HTTP {response.status_code}")

def parse_replay_data(replay_text):
    replay_info = {
        "format": None,
        "players": {},
        "teams": {"p1": {}, "p2": {}},
    }

    player_pokemon_map = {}

    match = re.search(r"\[.*\] .+: (.+?) vs. (.+?)(?= - Replays)", replay_text)
    if match:
        player1 = match.group(1).strip()
        player2 = match.group(2).strip()
        replay_info["players"]["p1"] = player1
        replay_info["players"]["p2"] = player2

    for line in replay_text.splitlines():
        if line.startswith("|poke|"):
            parts = line.split("|")
            player = parts[2]
            pokemon = parts[3].split(",")[0]
            replay_info["teams"][player][pokemon] = set()
        elif line.startswith("|switch|") or line.startswith("|drag|"):
            parts = line.split("|")
            identifier = parts[2]
            pokemon = parts[3].split(",")[0]
            player = identifier[:2]
            player_pokemon_map[identifier] = {
                "player": replay_info["players"].get(player, player),
                "pokemon": pokemon
            }
        elif line.startswith("|move|"):
            parts = line.split("|")
            user_id = parts[2]
            move = parts[3]

            user_info = player_pokemon_map.get(user_id, {"pokemon": user_id, "player": "Unknown"})
            player_key = "p1" if user_info["player"] == replay_info["players"].get("p1") else "p2"
            pokemon = user_info["pokemon"]

            if pokemon in replay_info["teams"][player_key]:
                if len(replay_info["teams"][player_key][pokemon]) < 4:
                    replay_info["teams"][player_key][pokemon].add(move)
        elif line.startswith("|tier|"):
            replay_info["format"] = line.split("|")[2]

    return replay_info


def save_to_excel(replay_info, replay_url, log_data):
    file_name = "pokemon_replay_data.xlsx"
    sprite_folder = "regular"

    if not os.path.exists(sprite_folder):
        raise Exception(f"Sprite folder '{sprite_folder}' not found.")

    # Custom Colors
    color_1 = PatternFill(start_color="90D5FF", end_color="90D5FF", fill_type="solid")  # #90D5FF
    color_2 = PatternFill(start_color="4BACC6", end_color="4BACC6", fill_type="solid")  # #57B9FF
    color_3 = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")  # #77B1D4
    color_4 = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")  # #517891

    # Bold font definition
    bold_font = Font(bold=True)

    try:
        workbook = load_workbook(file_name)
    except FileNotFoundError:
        workbook = Workbook()

    if "Matches" not in workbook.sheetnames:
        workbook.create_sheet("Matches")
    if "Teams & Moves" not in workbook.sheetnames:
        workbook.create_sheet("Teams & Moves")
    if "Overall Pokémon Usage" not in workbook.sheetnames:
        workbook.create_sheet("Overall Pokémon Usage")

    matches_sheet = workbook["Matches"]
    teams_sheet = workbook["Teams & Moves"]
    usage_sheet = workbook["Overall Pokémon Usage"]

    # Add headers to Overall Pokémon Usage sheet
    if usage_sheet.max_row == 1:
        usage_sheet.append(["Pokemon", "Total Usage", "Usage Percentage"])
        for col in range(1, usage_sheet.max_column + 1):
            usage_sheet.cell(row=1, column=col).fill = color_2

    # Add headers if missing for Matches sheet
    if matches_sheet.max_row == 1:
        matches_sheet.append(["Match ID", "Format", "Link", "Date", "Player 1", "Player 2", "Winner"])
        for col in range(1, matches_sheet.max_column + 1):
            matches_sheet.cell(row=1, column=col).fill = color_2

    # Add headers if missing for Teams & Moves sheet
    if teams_sheet.max_row == 1:
        teams_sheet.append(["Match ID", "Player", "Sprites", "Pokemon", "Move 1", "Move 2", "Move 3", "Move 4"])
        for col in range(1, teams_sheet.max_column + 1):
            teams_sheet.cell(row=1, column=col).fill = color_2

    # Apply color to entire column A (Match-ID) regardless of empty cells
    for row in teams_sheet.iter_rows(min_row=1, max_row=teams_sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.fill = color_2
            cell.border = None  # Remove borders

    # Apply color to entire column B (Player names) regardless of empty cells
    for row in teams_sheet.iter_rows(min_row=1, max_row=teams_sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.fill = color_2
            cell.border = None  # Remove borders

    # Apply color to entire column C (Sprites) regardless of empty cells
    for row in teams_sheet.iter_rows(min_row=1, max_row=teams_sheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.fill = color_2
            cell.border = None  # Remove borders

    match_id = matches_sheet.max_row
    match_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Append match data to Matches sheet
    winner = None
    for line in log_data.splitlines():
        if line.startswith("|win|"):
            winner = line.split("|")[2]
            break

    # Append match data to Matches sheet
    matches_sheet.append([
        f"Match {match_id}",
        replay_info["format"],
        replay_url,
        match_date,
        replay_info["players"].get("p1", "Unknown"),
        replay_info["players"].get("p2", "Unknown"),
        winner if winner else "Unknown"
    ])

    # Apply alternating colors for Player names in Teams & Moves sheet
    for row_idx, row in enumerate(teams_sheet.iter_rows(min_row=2, max_row=teams_sheet.max_row, min_col=2, max_col=2)):
        for cell in row:
            # Alternate colors color_3 and color_4
            if row_idx % 2 == 0:  # even row
                cell.fill = color_3
            else:  # odd row
                cell.fill = color_4
            cell.border = None  # Remove borders

    # Apply color to sprites column (C) in Teams & Moves sheet
    for row in teams_sheet.iter_rows(min_row=2, max_row=teams_sheet.max_row, min_col=3, max_col=3):
        for cell in row:
            cell.fill = color_2
            cell.border = None  # Remove borders

    # Add a separator row in Teams sheet
    separator_row = [f"Match {match_id}"] + ["+++++++++"] * 7
    teams_sheet.append(separator_row)

    # Make the separator row bold
    for col in range(1, len(separator_row) + 1):
        cell = teams_sheet.cell(row=teams_sheet.max_row, column=col)
        cell.font = Font(bold=True)

    # Keep track of Pokémon usage
    pokemon_usage = {}

    # Append team data to Teams sheet and track Pokémon usage
    for player_key, team in replay_info["teams"].items():
        player_name = replay_info["players"].get(player_key, "Unknown")
        teams_sheet.append([player_name])
        player_cell = teams_sheet.cell(row=teams_sheet.max_row, column=1)
        player_cell.fill = color_3
        player_cell.font = bold_font  # Bold player name
        player_cell.border = None  # Remove borders

        for pokemon, moves in team.items():
            sprite_path = os.path.join(sprite_folder, f"{pokemon}.png")
            if os.path.exists(sprite_path):
                img = Image(sprite_path)
                img.height = 40
                img.width = 40

                sprite_cell = teams_sheet.cell(row=teams_sheet.max_row + 1, column=3)
                sprite_cell.fill = color_2
                teams_sheet.add_image(img, sprite_cell.coordinate)

            moves_list = list(moves) + [""] * (4 - len(moves))
            team_row = ["", "", "", pokemon] + moves_list
            teams_sheet.append(team_row)

            # Count Pokémon usage
            pokemon_usage[pokemon] = pokemon_usage.get(pokemon, 0) + 1

    # Apply alternating row color for Pokémon, moves columns (D-H) in Teams sheet
    for row_idx, row in enumerate(teams_sheet.iter_rows(min_row=2, max_row=teams_sheet.max_row, min_col=4, max_col=8)):
        for cell in row:
            if row_idx % 2 == 0:  # even row
                cell.fill = color_3
            else:  # odd row
                cell.fill = color_4
            cell.border = None  # Remove borders

    # **New Code** to handle merging and sorting Pokémon usage data
    # Extract existing data from the sheet (excluding headers)
    existing_data = {}
# Read existing data from the sheet
    for row in usage_sheet.iter_rows(min_row=2, max_row=usage_sheet.max_row, min_col=1, max_col=3, values_only=True):
        if row[0]:  # Check if the Pokémon name exists
            name = row[0]
            count = int(row[1])
            percentage = float(row[2].replace('%', ''))
            existing_data[name] = {"count": count, "percentage": percentage}

    # Add new data to the existing data, updating counts
    for name, count in pokemon_usage.items():
        if name in existing_data:
            existing_data[name]["count"] += count  # Update count
        else:
            existing_data[name] = {"count": count, "percentage": 0}  # Add new Pokémon with count

    # Calculate the total replays
    total_replays = len(matches_sheet["A"]) - 1  # Adjust this to reflect your actual total replay count
    total_teams = total_replays * 2  # Each replay has 2 teams

    # Recalculate percentages for all Pokémon using the total teams
    for data in existing_data.values():
        data["percentage"] = (data["count"] / total_teams) * 100  # Calculate percentage based on total teams

    # Sort data by percentage in descending order
    sorted_data = sorted(
        existing_data.items(),
        key=lambda x: x[1]["percentage"],
        reverse=True
    )

    # Clear existing rows in the sheet after the header
    for row in usage_sheet.iter_rows(min_row=2, max_row=usage_sheet.max_row):
        for cell in row:
            cell.value = None

    # Write sorted data back to the sheet
    for i, (name, stats) in enumerate(sorted_data, start=2):
        usage_sheet.cell(row=i, column=1, value=name)  # Pokémon Name
        usage_sheet.cell(row=i, column=2, value=stats["count"])  # Count
        usage_sheet.cell(row=i, column=3, value=f"{stats['percentage']:.2f}%")  # Percentage


    workbook.save(file_name)

def process_replay(link_entry):
    replay_url = link_entry.get().strip()
    if not replay_url:
        messagebox.showerror("Error", "Please enter a valid link.") 
        return

    try:
        replay_text = fetch_replay_data(replay_url)
        parsed_data = parse_replay_data(replay_text)
        save_to_excel(parsed_data, replay_url,log_data=replay_text)
        messagebox.showinfo("Success", "Replay data saved successfully to Excel.")
        link_entry.delete(0, "end")
    except Exception as e:
        messagebox.showerror("Error", str(e))

def open_spreadsheet():
    try:
        os.startfile("pokemon_replay_data.xlsx")
    except FileNotFoundError:
        print("Error: File 'pokemon_replay_data.xlsx' not found.")


def inspect_players():
    try:
        os.startfile("inspector.exe")
    except FileNotFoundError:
        print("Error: File 'inspector.exe' not found.")



def create_gui():
    root = Tk()
    root.title("Pokémon Replay Processor")

    frame = Frame(root, padx=20, pady=20)
    frame.pack()

    Label(frame, text="Empire Scout", font=("Arial", 16, "bold"), fg="black").grid(row=0, column=0, columnspan=2, pady=(0, 20))

    Label(frame, text="Enter Replay Link:", font=("Arial", 12)).grid(row=1, column=0, sticky="e", padx=10, pady=5)
    link_entry = Entry(frame, width=50, font=("Arial", 12))
    link_entry.grid(row=1, column=1, pady=5)

    button_width = 20

    Button(frame, text="Submit", command=lambda: process_replay(link_entry), font=("Arial", 12), bg="black", fg="white", width=button_width).grid(row=2, column=0, columnspan=2, pady=(10, 10))

    Button(frame, text="Open Spreadsheet", command=open_spreadsheet, font=("Arial", 12), bg="blue", fg="white", width=button_width).grid(row=3, column=0, columnspan=2, pady=(5, 5))

    Button(frame, text="Inspect Players", command=inspect_players, font=("Arial", 12), bg="green", fg="white", width=button_width).grid(row=4, column=0, columnspan=2, pady=(5, 20))

    instructions = ("Instructions:\n"
                    "1. Make sure the link is correct.\n"
                    "2. Close the Excel sheet before processing.\n"
                    "3. Process one link at a time.")
    Label(frame, text=instructions, font=("Arial", 10), justify="left", fg="darkred").grid(row=5, column=0, columnspan=2, pady=10)

    footer_text = (
        "Created by subr0za\n"
        "For bug reports, contact me on Discord."
    )
    Label(root, text=footer_text, font=("Arial", 10, "italic"), fg="gray").pack(side="bottom", pady=10)

    root.mainloop()

def inspect_players():
    messagebox.showinfo("Coming Soon", "Player inspection feature is coming soon!")


if __name__ == "__main__":
    create_gui()
